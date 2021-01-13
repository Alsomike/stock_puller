import openpyxl as xl
import time

trial_stocks = ['stock1', 'stock2', 'stock3', 'stock4', 'stock5']
trial_prices = ['price1', 'price2', 'price3', 'price4', 'price5']

def Add_To_Excel():
	now = time.strftime("%x") #establishes variable now that gives us the current date

	SR_wb = xl.load_workbook('\\Users\\mpell\\OneDrive\\Documents\\PythonProjects\\Git\\Stock_Records.xlsx') #loads excel file using where it lies in the directory

	SR_Sheet = SR_wb.create_sheet() #creates a new sheet within the SR_wb

	SR_Sheet['A1'].value = now #adds the current date to the first cell
	SR_Sheet['C1'].value = 'Ticker'
	SR_Sheet['D1'].value = 'Price'

	for stocks in range(0, len(trial_stocks)):
		SR_Sheet.cell(row = stocks+2,column = 3).value=trial_stocks[stocks]
	for prices in range(0, len(trial_stocks)):
		SR_Sheet.cell(row = prices+2,column = 4).value=trial_prices[prices]

	SR_wb.save('Stock_Records.xlsx') #saves changes made to excel file

Add_To_Excel()