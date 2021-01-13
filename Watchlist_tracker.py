import requests
import bs4
import time
import openpyxl as xl

UserInput = 0
StockPrice = []
StockTicker = []
def getStock(i):
	res = requests.get(i)
	soup = bs4.BeautifulSoup(res.text, features="html.parser")
	HTMLSelector = soup.select('#quote-header-info > div.My\(6px\).Pos\(r\).smartphone_Mt\(6px\) > div.D\(ib\).Va\(m\).Maw\(65\%\).Ov\(h\) > div > span.Trsdu\(0\.3s\).Fw\(b\).Fz\(36px\).Mb\(-4px\).D\(ib\)')
	stockprint = HTMLSelector[0].text.strip()
	StockPrice.append(stockprint)
	print('$' + stockprint)

def Add_To_Excel():
	now = time.strftime("%x") #establishes variable now that gives us the current date

	SR_wb = xl.load_workbook('\\Users\\mpell\\OneDrive\\Documents\\PythonProjects\\Git\\Stock_Records.xlsx') #loads excel file using where it lies in the directory

	SR_Sheet = SR_wb.create_sheet() #creates a new sheet within the SR_wb

	SR_Sheet['A1'].value = now #adds the current date to the first cell
	SR_Sheet['C1'].value = 'Ticker'
	SR_Sheet['D1'].value = 'Price'

	for stocks in range(0, len(StockTicker)):
		SR_Sheet.cell(row = stocks+2,column = 3).value=StockTicker[stocks]
	for prices in range(0, len(StockPrice)):
		SR_Sheet.cell(row = prices+2,column = 4).value=StockPrice[prices]

	SR_wb.save('Stock_Records.xlsx') #saves changes made to excel file
while True:
	if UserInput != '1':
		try:
			UserInput = input("what is the ticker? ")
			StockTicker.append(UserInput)
			getStock('https://finance.yahoo.com/quote/' + UserInput + '?p=TSLA&.tsrc=fin-srch')
			print(StockTicker, StockPrice)
		except:
			break
Add_To_Excel()
print('break')